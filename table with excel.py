import sys
import asyncio
import logging
import json
from pathlib import Path
from PyQt5.QtWidgets import (QApplication, QMainWindow, QTableWidget, QTableWidgetItem, 
                             QPushButton, QVBoxLayout, QWidget, QHBoxLayout, QMessageBox,
                             QLabel, QHeaderView, QTabWidget, QComboBox, QFileDialog,
                             QProgressDialog, QGroupBox, QGridLayout)
from PyQt5.QtCore import Qt, QTimer, QThread, pyqtSignal
from PyQt5.QtGui import QFont
from database import Database
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

GROUPS = [
    "VR/AR приложения",
    "3D моделирование",
    "Инженерная графика",
    "Основы графического дизайна",
    "Пайтон для начинающих",
    "Архитектура и дизайн",
    "Интеллект-школа. История.9-11 классы",
    "Интеллект школа. Русский язык. (7-8 классы)",
    "Концепт-арт: Создание 3D-персонажа",
    "Основы веб-программирования",
]

DAYS_OF_WEEK = [
    "понедельник",
    "вторник",
    "среда",
    "четверг",
    "пятница",
    "суббота",
    "воскресенье",
]

# Класс для асинхронной загрузки Excel
class ExcelLoaderThread(QThread):
    finished = pyqtSignal(list, list, str)  # users_data, schedule_data, report
    error = pyqtSignal(str)
    progress = pyqtSignal(int)
    
    def __init__(self, file_path, file_type):
        super().__init__()
        self.file_path = file_path
        self.file_type = file_type  # 'users' or 'schedule' or 'both'
    
    def run(self):
        try:
            users_data = []
            schedule_data = []
            report = []
            
            df = pd.read_excel(self.file_path)
            self.progress.emit(20)
            
            if self.file_type in ['users', 'both']:
                # Парсим учеников
                if 'user_id' in df.columns and 'full_name' in df.columns and 'group' in df.columns:
                    for idx, row in df.iterrows():
                        try:
                            user_id = int(row['user_id'])
                            full_name = str(row['full_name']).strip()
                            group = str(row['group']).strip()
                            
                            if group not in GROUPS:
                                report.append(f"⚠️ Строка {idx + 2}: группа '{group}' не найдена в списке")
                                continue
                            
                            name_parts = full_name.split()
                            first_name = name_parts[1] if len(name_parts) > 1 else name_parts[0]
                            last_name = name_parts[0] if name_parts else ''
                            
                            users_data.append((user_id, group, full_name, '', first_name, last_name))
                            report.append(f"✅ Добавлен ученик: {full_name} (ID: {user_id}, группа: {group})")
                        except Exception as e:
                            report.append(f"❌ Ошибка в строке {idx + 2}: {str(e)}")
            
            self.progress.emit(50)
            
            if self.file_type in ['schedule', 'both']:
                # Парсим расписание
                if 'group' in df.columns and 'day' in df.columns and 'time' in df.columns:
                    for idx, row in df.iterrows():
                        try:
                            group = str(row['group']).strip()
                            day = str(row['day']).strip().lower()
                            time = str(row['time']).strip()
                            
                            if group not in GROUPS:
                                report.append(f"⚠️ Строка {idx + 2}: группа '{group}' не найдена в списке")
                                continue
                            
                            if day not in DAYS_OF_WEEK:
                                report.append(f"⚠️ Строка {idx + 2}: день '{day}' не найден в списке")
                                continue
                            
                            schedule_data.append((group, day, time))
                            report.append(f"✅ Добавлено занятие: {group} - {day} - {time}")
                        except Exception as e:
                            report.append(f"❌ Ошибка в строке {idx + 2}: {str(e)}")
            
            self.progress.emit(100)
            self.finished.emit(users_data, schedule_data, "\n".join(report))
            
        except Exception as e:
            self.error.emit(str(e))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Административная панель - Управление данными")
        self.setGeometry(100, 100, 1400, 900)

        # Инициализация базы данных
        self.db = None
        self.db_initialized = False

        # Путь к файлу сохранения
        self.save_file_users = Path('data/table_users.json')
        self.save_file_schedule = Path('data/table_schedule.json')
        for file in [self.save_file_users, self.save_file_schedule]:
            file.parent.mkdir(parents=True, exist_ok=True)

        # Путь к шаблонам Excel
        self.templates_dir = Path('templates')
        self.templates_dir.mkdir(parents=True, exist_ok=True)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QVBoxLayout(central_widget)

        # Информационная метка
        info_label = QLabel(
            "📋 Работа в три этапа:\n"
            "1. Вкладка 'Учащиеся': добавьте всех учащихся (User ID, ФИО, Группа)\n"
            "2. Вкладка 'Расписание': добавьте расписание (Группа, День недели, Время)\n"
            "3. Вкладка 'Статистика': просмотр статистики по группам\n\n"
            "💡 Можно загрузить данные из Excel (кнопка 'Загрузить из Excel')"
        )
        info_label.setStyleSheet("padding: 10px; background-color: #e7f3ff; border-radius: 5px; font-size: 12px;")
        main_layout.addWidget(info_label)

        # Создаем вкладки
        self.tabs = QTabWidget()
        
        # Вкладка 1: Учащиеся
        self.setup_users_tab()
        
        # Вкладка 2: Расписание
        self.setup_schedule_tab()
        
        # Вкладка 3: Статистика
        self.setup_stats_tab()
        
        main_layout.addWidget(self.tabs)

        # Панель инструментов
        toolbar_layout = QHBoxLayout()
        
        # Кнопки работы с Excel
        excel_group = QGroupBox("Работа с Excel")
        excel_layout = QHBoxLayout()
        
        self.btn_load_excel = QPushButton("📥 Загрузить из Excel")
        self.btn_load_excel.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #138496;
            }
        """)
        self.btn_load_excel.clicked.connect(self.load_from_excel)
        
        self.btn_export_excel = QPushButton("📤 Экспорт в Excel")
        self.btn_export_excel.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        self.btn_export_excel.clicked.connect(self.export_to_excel)
        
        self.btn_template = QPushButton("📋 Шаблоны Excel")
        self.btn_template.setStyleSheet("""
            QPushButton {
                background-color: #ffc107;
                color: black;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #e0a800;
            }
        """)
        self.btn_template.clicked.connect(self.create_templates)
        
        excel_layout.addWidget(self.btn_load_excel)
        excel_layout.addWidget(self.btn_export_excel)
        excel_layout.addWidget(self.btn_template)
        excel_group.setLayout(excel_layout)
        toolbar_layout.addWidget(excel_group)
        
        # Кнопки управления БД
        db_group = QGroupBox("База данных")
        db_layout = QHBoxLayout()
        
        self.button_save = QPushButton("💾 Сохранить")
        self.button_save.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        self.button_save.clicked.connect(self.save_all_data)

        self.button_add = QPushButton("🚀 Загрузить в БД")
        self.button_add.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        self.button_add.clicked.connect(self.load_data_to_database)
        
        db_layout.addWidget(self.button_save)
        db_layout.addWidget(self.button_add)
        db_group.setLayout(db_layout)
        toolbar_layout.addWidget(db_group)
        
        toolbar_layout.addStretch()
        main_layout.addLayout(toolbar_layout)

        # Статусная панель
        status_panel = QHBoxLayout()
        
        self.status_label = QLabel("✅ Готов к работе")
        self.status_label.setStyleSheet("padding: 8px; background-color: #e7f3ff; border-radius: 5px;")
        status_panel.addWidget(self.status_label)
        
        self.stats_label = QLabel("👥 Учеников: 0 | 📅 Занятий: 0")
        self.stats_label.setStyleSheet("padding: 8px; background-color: #f8f9fa; border-radius: 5px;")
        status_panel.addWidget(self.stats_label)
        
        status_panel.addStretch()
        main_layout.addLayout(status_panel)

        # Таймер для автосохранения
        self.auto_save_timer = QTimer()
        self.auto_save_timer.setSingleShot(True)
        self.auto_save_timer.timeout.connect(self.save_all_data)
        self.auto_save_timer.setInterval(1000)

        # Загружаем сохраненные данные при запуске
        QTimer.singleShot(100, self.load_all_data)

        # Подключаем сигналы изменения данных
        self.users_table.itemChanged.connect(self.trigger_auto_save)
        self.schedule_table.itemChanged.connect(self.trigger_auto_save)

    def setup_users_tab(self):
        """Настройка вкладки учащихся"""
        self.users_tab = QWidget()
        users_layout = QVBoxLayout(self.users_tab)
        
        # Информационная панель
        users_info = QLabel(
            "👥 Добавление учащихся\n"
            "• User ID - уникальный идентификатор (число)\n"
            "• ФИО - полное имя ученика\n"
            "• Группа - выберите из списка"
        )
        users_info.setStyleSheet("padding: 8px; background-color: #f0f0f0; border-radius: 5px;")
        users_layout.addWidget(users_info)
        
        # Таблица учащихся
        self.users_table = QTableWidget()
        self.users_table.setRowCount(50)
        self.users_table.setColumnCount(3)
        self.users_table.setHorizontalHeaderLabels(["User ID", "ФИО", "Группа"])
        self.users_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.users_table.setAlternatingRowColors(True)
        self.users_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #d0d0d0;
                background-color: white;
                font-size: 12px;
            }
            QTableWidget::item {
                padding: 8px;
            }
            QTableWidget::item:selected {
                background-color: #cce5ff;
            }
        """)
        users_layout.addWidget(self.users_table)
        self._ensure_group_widgets_for_all_rows()
        
        # Кнопки управления
        users_buttons = QHBoxLayout()
        
        self.users_add_row = QPushButton("➕ Добавить строку")
        self.users_add_row.setStyleSheet("background-color: #17a2b8; color: white; padding: 8px; border-radius: 5px;")
        self.users_add_row.clicked.connect(lambda: self.add_row(self.users_table))
        
        self.users_delete = QPushButton("🗑️ Удалить строку")
        self.users_delete.setStyleSheet("background-color: #dc3545; color: white; padding: 8px; border-radius: 5px;")
        self.users_delete.clicked.connect(lambda: self.delete_selected_row(self.users_table))
        
        self.users_clear = QPushButton("🧹 Очистить")
        self.users_clear.setStyleSheet("background-color: #ffc107; color: black; padding: 8px; border-radius: 5px;")
        self.users_clear.clicked.connect(lambda: self.clear_table(self.users_table))
        
        users_buttons.addWidget(self.users_add_row)
        users_buttons.addWidget(self.users_delete)
        users_buttons.addWidget(self.users_clear)
        users_buttons.addStretch()
        users_layout.addLayout(users_buttons)
        
        self.tabs.addTab(self.users_tab, "👥 Учащиеся")

    def setup_schedule_tab(self):
        """Настройка вкладки расписания"""
        self.schedule_tab = QWidget()
        schedule_layout = QVBoxLayout(self.schedule_tab)
        
        # Информационная панель
        schedule_info = QLabel(
            "📅 Добавление расписания\n"
            "• Группа - выберите из списка\n"
            "• День недели - выберите день\n"
            "• Время - формат ЧЧ:ММ (например, 15:30)"
        )
        schedule_info.setStyleSheet("padding: 8px; background-color: #f0f0f0; border-radius: 5px;")
        schedule_layout.addWidget(schedule_info)
        
        # Таблица расписания
        self.schedule_table = QTableWidget()
        self.schedule_table.setRowCount(100)
        self.schedule_table.setColumnCount(3)
        self.schedule_table.setHorizontalHeaderLabels(["Группа", "День недели", "Время"])
        self.schedule_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.schedule_table.setAlternatingRowColors(True)
        self.schedule_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #d0d0d0;
                background-color: white;
                font-size: 12px;
            }
            QTableWidget::item {
                padding: 8px;
            }
        """)
        schedule_layout.addWidget(self.schedule_table)
        self._ensure_schedule_widgets_for_all_rows()

        # Кнопки управления
        schedule_buttons = QHBoxLayout()
        
        self.schedule_add_row = QPushButton("➕ Добавить строку")
        self.schedule_add_row.setStyleSheet("background-color: #17a2b8; color: white; padding: 8px; border-radius: 5px;")
        self.schedule_add_row.clicked.connect(lambda: self.add_row(self.schedule_table))
        
        self.schedule_delete = QPushButton("🗑️ Удалить строку")
        self.schedule_delete.setStyleSheet("background-color: #dc3545; color: white; padding: 8px; border-radius: 5px;")
        self.schedule_delete.clicked.connect(lambda: self.delete_selected_row(self.schedule_table))
        
        self.schedule_clear = QPushButton("🧹 Очистить")
        self.schedule_clear.setStyleSheet("background-color: #ffc107; color: black; padding: 8px; border-radius: 5px;")
        self.schedule_clear.clicked.connect(lambda: self.clear_table(self.schedule_table))
        
        schedule_buttons.addWidget(self.schedule_add_row)
        schedule_buttons.addWidget(self.schedule_delete)
        schedule_buttons.addWidget(self.schedule_clear)
        schedule_buttons.addStretch()
        schedule_layout.addLayout(schedule_buttons)
        
        self.tabs.addTab(self.schedule_tab, "📅 Расписание")

    def setup_stats_tab(self):
        """Настройка вкладки статистики"""
        self.stats_tab = QWidget()
        stats_layout = QVBoxLayout(self.stats_tab)
        
        # Заголовок
        stats_title = QLabel("📊 Статистика по группам")
        stats_title.setStyleSheet("font-size: 18px; font-weight: bold; padding: 10px;")
        stats_layout.addWidget(stats_title)
        
        # Статистика по группам
        self.stats_group = QGroupBox("Распределение учащихся")
        self.stats_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        stats_grid = QGridLayout()
        
        # Заголовки
        headers = ["Группа", "Кол-во учеников", "Кол-во занятий", "Статус"]
        for col, header in enumerate(headers):
            label = QLabel(header)
            label.setStyleSheet("font-weight: bold; padding: 5px; background-color: #f0f0f0;")
            stats_grid.addWidget(label, 0, col)
        
        self.stats_labels = {}
        for row, group in enumerate(GROUPS, start=1):
            # Название группы
            group_label = QLabel(group)
            group_label.setWordWrap(True)
            stats_grid.addWidget(group_label, row, 0)
            
            # Количество учеников
            students_label = QLabel("0")
            students_label.setAlignment(Qt.AlignCenter)
            stats_grid.addWidget(students_label, row, 1)
            
            # Количество занятий
            lessons_label = QLabel("0")
            lessons_label.setAlignment(Qt.AlignCenter)
            stats_grid.addWidget(lessons_label, row, 2)
            
            # Статус
            status_label = QLabel("❌ Нет")
            status_label.setAlignment(Qt.AlignCenter)
            stats_grid.addWidget(status_label, row, 3)
            
            self.stats_labels[group] = {
                'students': students_label,
                'lessons': lessons_label,
                'status': status_label
            }
        
        self.stats_group.setLayout(stats_grid)
        stats_layout.addWidget(self.stats_group)
        
        # Общая статистика
        total_group = QGroupBox("Общие показатели")
        total_layout = QGridLayout()
        
        self.total_students_label = QLabel("Всего учеников: 0")
        self.total_students_label.setStyleSheet("font-size: 14px; padding: 5px;")
        total_layout.addWidget(self.total_students_label, 0, 0)
        
        self.total_lessons_label = QLabel("Всего занятий: 0")
        self.total_lessons_label.setStyleSheet("font-size: 14px; padding: 5px;")
        total_layout.addWidget(self.total_lessons_label, 0, 1)
        
        self.total_groups_students = QLabel("Групп с учениками: 0")
        self.total_groups_students.setStyleSheet("font-size: 14px; padding: 5px;")
        total_layout.addWidget(self.total_groups_students, 1, 0)
        
        self.total_groups_schedule = QLabel("Групп с занятиями: 0")
        self.total_groups_schedule.setStyleSheet("font-size: 14px; padding: 5px;")
        total_layout.addWidget(self.total_groups_schedule, 1, 1)
        
        total_group.setLayout(total_layout)
        stats_layout.addWidget(total_group)
        
        # Кнопка обновления
        self.btn_refresh_stats = QPushButton("🔄 Обновить статистику")
        self.btn_refresh_stats.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        self.btn_refresh_stats.clicked.connect(self.update_statistics)
        stats_layout.addWidget(self.btn_refresh_stats)
        
        stats_layout.addStretch()
        
        self.tabs.addTab(self.stats_tab, "📊 Статистика")

    def load_from_excel(self):
        """Загрузка данных из Excel"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            "Выберите файл Excel", 
            str(Path.home()), 
            "Excel files (*.xlsx *.xls)"
        )
        
        if not file_path:
            return
        
        # Спрашиваем тип данных
        msg = QMessageBox()
        msg.setWindowTitle("Тип данных")
        msg.setText("Выберите тип данных для загрузки:")
        btn_users = msg.addButton("Только ученики", QMessageBox.ActionRole)
        btn_schedule = msg.addButton("Только расписание", QMessageBox.ActionRole)
        btn_both = msg.addButton("Оба", QMessageBox.ActionRole)
        btn_cancel = msg.addButton("Отмена", QMessageBox.RejectRole)
        
        msg.exec_()
        
        if msg.clickedButton() == btn_cancel:
            return
        elif msg.clickedButton() == btn_users:
            file_type = 'users'
        elif msg.clickedButton() == btn_schedule:
            file_type = 'schedule'
        else:
            file_type = 'both'
        
        # Показываем прогресс
        self.progress = QProgressDialog("Загрузка данных...", "Отмена", 0, 100, self)
        self.progress.setWindowModality(Qt.WindowModal)
        self.progress.show()
        
        # Запускаем загрузку в отдельном потоке
        self.loader_thread = ExcelLoaderThread(file_path, file_type)
        self.loader_thread.finished.connect(self.on_excel_loaded)
        self.loader_thread.error.connect(self.on_excel_error)
        self.loader_thread.progress.connect(self.progress.setValue)
        self.loader_thread.start()

    def on_excel_loaded(self, users_data, schedule_data, report):
        """Обработка завершения загрузки из Excel"""
        self.progress.close()
        
        # Заполняем таблицы
        if users_data:
            self.users_table.setRowCount(len(users_data))
            self._ensure_group_widgets_for_all_rows()
            
            for row, (user_id, group, full_name, username, first_name, last_name) in enumerate(users_data):
                self.users_table.setItem(row, 0, QTableWidgetItem(str(user_id)))
                self.users_table.setItem(row, 1, QTableWidgetItem(full_name))
                combo = self.users_table.cellWidget(row, 2)
                if isinstance(combo, QComboBox):
                    combo.setCurrentText(group)
        
        if schedule_data:
            self.schedule_table.setRowCount(len(schedule_data))
            self._ensure_schedule_widgets_for_all_rows()
            
            for row, (group, day, time) in enumerate(schedule_data):
                combo_group = self.schedule_table.cellWidget(row, 0)
                combo_day = self.schedule_table.cellWidget(row, 1)
                if isinstance(combo_group, QComboBox):
                    combo_group.setCurrentText(group)
                if isinstance(combo_day, QComboBox):
                    combo_day.setCurrentText(day.capitalize())
                self.schedule_table.setItem(row, 2, QTableWidgetItem(time))
        
        # Показываем отчет
        report_dialog = QMessageBox(self)
        report_dialog.setWindowTitle("Отчет о загрузке")
        report_dialog.setText("Загрузка из Excel завершена")
        report_dialog.setDetailedText(report)
        report_dialog.setIcon(QMessageBox.Information)
        report_dialog.exec_()
        
        # Обновляем статистику
        self.update_statistics()
        self.status_label.setText("✅ Данные загружены из Excel")

    def on_excel_error(self, error_msg):
        """Обработка ошибки загрузки из Excel"""
        self.progress.close()
        QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл:\n{error_msg}")

    def export_to_excel(self):
        """Экспорт данных в Excel"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Сохранить как Excel", 
            f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            wb = openpyxl.Workbook()
            
            # Лист с учениками
            ws_users = wb.active
            ws_users.title = "Учащиеся"
            
            headers_users = ["User ID", "ФИО", "Группа"]
            for col, header in enumerate(headers_users, 1):
                cell = ws_users.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            
            row_idx = 2
            for row in range(self.users_table.rowCount()):
                user_id = self._get_table_cell_text(self.users_table, row, 0)
                full_name = self._get_table_cell_text(self.users_table, row, 1)
                group = self._get_table_cell_text(self.users_table, row, 2)
                
                if user_id and full_name and group:
                    ws_users.cell(row=row_idx, column=1, value=int(user_id) if user_id.isdigit() else user_id)
                    ws_users.cell(row=row_idx, column=2, value=full_name)
                    ws_users.cell(row=row_idx, column=3, value=group)
                    row_idx += 1
            
            # Лист с расписанием
            ws_schedule = wb.create_sheet("Расписание")
            
            headers_schedule = ["Группа", "День недели", "Время"]
            for col, header in enumerate(headers_schedule, 1):
                cell = ws_schedule.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            
            row_idx = 2
            for row in range(self.schedule_table.rowCount()):
                group = self._get_table_cell_text(self.schedule_table, row, 0)
                day = self._get_table_cell_text(self.schedule_table, row, 1)
                time = self._get_table_cell_text(self.schedule_table, row, 2)
                
                if group and day and time:
                    ws_schedule.cell(row=row_idx, column=1, value=group)
                    ws_schedule.cell(row=row_idx, column=2, value=day)
                    ws_schedule.cell(row=row_idx, column=3, value=time)
                    row_idx += 1
            
            # Лист со статистикой
            ws_stats = wb.create_sheet("Статистика")
            
            stats_headers = ["Группа", "Учеников", "Занятий"]
            for col, header in enumerate(stats_headers, 1):
                cell = ws_stats.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
            
            users_data = self.get_users_data()
            schedule_data = self.get_schedule_data()
            
            group_counts = {}
            for _, group_name, _, _, _, _ in users_data:
                group_counts[group_name] = group_counts.get(group_name, 0) + 1
            
            schedule_by_group = {}
            for group_name, _, _ in schedule_data:
                schedule_by_group[group_name] = schedule_by_group.get(group_name, 0) + 1
            
            row_idx = 2
            for group in GROUPS:
                students = group_counts.get(group, 0)
                lessons = schedule_by_group.get(group, 0)
                if students > 0 or lessons > 0:
                    ws_stats.cell(row=row_idx, column=1, value=group)
                    ws_stats.cell(row=row_idx, column=2, value=students)
                    ws_stats.cell(row=row_idx, column=3, value=lessons)
                    row_idx += 1
            
            wb.save(file_path)
            
            QMessageBox.information(self, "Успех", f"Данные экспортированы в файл:\n{file_path}")
            self.status_label.setText(f"✅ Данные экспортированы в {file_path}")
            
        except Exception as e:
            logger.error(f"Ошибка при экспорте: {e}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать данные:\n{str(e)}")

    def create_templates(self):
        """Создание шаблонов Excel"""
        try:
            # Шаблон для учеников
            wb_users = openpyxl.Workbook()
            ws = wb_users.active
            ws.title = "Учащиеся"
            
            headers = ["user_id", "full_name", "group"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            
            example_data = [
                [1001, "Иванов Иван", "VR/AR приложения"],
                [1002, "Петрова Мария", "3D моделирование"],
            ]
            for row, data in enumerate(example_data, 2):
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            users_template = self.templates_dir / "template_users.xlsx"
            wb_users.save(users_template)
            
            # Шаблон для расписания
            wb_schedule = openpyxl.Workbook()
            ws = wb_schedule.active
            ws.title = "Расписание"
            
            headers = ["group", "day", "time"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            
            example_data = [
                ["VR/AR приложения", "понедельник", "15:00"],
                ["3D моделирование", "среда", "16:30"],
            ]
            for row, data in enumerate(example_data, 2):
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            schedule_template = self.templates_dir / "template_schedule.xlsx"
            wb_schedule.save(schedule_template)
            
            QMessageBox.information(
                self, 
                "Успех", 
                f"Шаблоны созданы в папке:\n{self.templates_dir.absolute()}\n\n"
                f"• template_users.xlsx - для учеников\n"
                f"• template_schedule.xlsx - для расписания"
            )
            
        except Exception as e:
            logger.error(f"Ошибка при создании шаблонов: {e}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось создать шаблоны:\n{str(e)}")

    def update_statistics(self):
        """Обновление статистики"""
        users_data = self.get_users_data()
        schedule_data = self.get_schedule_data()
        
        group_counts = {}
        for _, group_name, _, _, _, _ in users_data:
            group_counts[group_name] = group_counts.get(group_name, 0) + 1
        
        schedule_by_group = {}
        for group_name, _, _ in schedule_data:
            schedule_by_group[group_name] = schedule_by_group.get(group_name, 0) + 1
        
        for group in GROUPS:
            if group in self.stats_labels:
                students = group_counts.get(group, 0)
                lessons = schedule_by_group.get(group, 0)
                
                self.stats_labels[group]['students'].setText(str(students))
                self.stats_labels[group]['lessons'].setText(str(lessons))
                
                if students > 0 or lessons > 0:
                    self.stats_labels[group]['status'].setText("✅ Да")
                    self.stats_labels[group]['status'].setStyleSheet("color: green; font-weight: bold;")
                else:
                    self.stats_labels[group]['status'].setText("❌ Нет")
                    self.stats_labels[group]['status'].setStyleSheet("color: red;")
        
        total_students = len(users_data)
        total_lessons = len(schedule_data)
        groups_with_students = len([g for g in GROUPS if group_counts.get(g, 0) > 0])
        groups_with_schedule = len([g for g in GROUPS if schedule_by_group.get(g, 0) > 0])
        
        self.total_students_label.setText(f"Всего учеников: {total_students}")
        self.total_lessons_label.setText(f"Всего занятий: {total_lessons}")
        self.total_groups_students.setText(f"Групп с учениками: {groups_with_students}")
        self.total_groups_schedule.setText(f"Групп с занятиями: {groups_with_schedule}")
        
        self.stats_label.setText(f"👥 Учеников: {total_students} | 📅 Занятий: {total_lessons}")

    def trigger_auto_save(self):
        """Запуск автосохранения"""
        self.auto_save_timer.start()

    def add_row(self, table):
        """Добавление строки"""
        current_row = table.rowCount()
        table.insertRow(current_row)
        
        if table == self.users_table:
            self._ensure_group_widget(current_row)
        else:
            self._ensure_schedule_row_widgets(current_row)
        
        self.status_label.setText(f"Добавлена строка {current_row + 1}")

    def delete_selected_row(self, table):
        """Удаление выбранной строки"""
        current_row = table.currentRow()
        if current_row >= 0:
            table.removeRow(current_row)
            self.status_label.setText(f"Строка {current_row + 1} удалена")
            self.trigger_auto_save()
        else:
            QMessageBox.warning(self, "Предупреждение", "Выберите строку для удаления")

    def clear_table(self, table):
        """Очистка таблицы"""
        reply = QMessageBox.question(self, "Подтверждение", 
                                    "Вы уверены, что хотите очистить таблицу?",
                                    QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            table.setRowCount(50 if table == self.users_table else 100)
            for row in range(table.rowCount()):
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    if item:
                        item.setText("")
            
            if table == self.users_table:
                self._ensure_group_widgets_for_all_rows()
                for row in range(table.rowCount()):
                    combo = self.users_table.cellWidget(row, 2)
                    if isinstance(combo, QComboBox):
                        combo.setCurrentIndex(0)
            else:
                self._ensure_schedule_widgets_for_all_rows()
                for row in range(table.rowCount()):
                    for col in (0, 1):
                        combo = self.schedule_table.cellWidget(row, col)
                        if isinstance(combo, QComboBox):
                            combo.setCurrentIndex(0)
                    item = self.schedule_table.item(row, 2)
                    if item:
                        item.setText("")
            
            self.status_label.setText("Таблица очищена")
            self.trigger_auto_save()

    def _ensure_group_widgets_for_all_rows(self):
        for row in range(self.users_table.rowCount()):
            self._ensure_group_widget(row)

    def _ensure_group_widget(self, row):
        if not isinstance(self.users_table.cellWidget(row, 2), QComboBox):
            combo = QComboBox()
            combo.addItem("")
            combo.addItems(GROUPS)
            combo.currentTextChanged.connect(lambda: self.trigger_auto_save())
            self.users_table.setCellWidget(row, 2, combo)

    def _ensure_schedule_widgets_for_all_rows(self):
        for row in range(self.schedule_table.rowCount()):
            self._ensure_schedule_row_widgets(row)

    def _ensure_schedule_row_widgets(self, row):
        for col, items in [(0, GROUPS), (1, DAYS_OF_WEEK)]:
            if not isinstance(self.schedule_table.cellWidget(row, col), QComboBox):
                combo = QComboBox()
                combo.addItem("")
                combo.addItems(items)
                combo.currentTextChanged.connect(lambda: self.trigger_auto_save())
                self.schedule_table.setCellWidget(row, col, combo)
        
        if self.schedule_table.item(row, 2) is None:
            self.schedule_table.setItem(row, 2, QTableWidgetItem(""))

    def _get_table_cell_text(self, table, row, col):
        widget = table.cellWidget(row, col)
        if isinstance(widget, QComboBox):
            return widget.currentText().strip()
        item = table.item(row, col)
        return item.text().strip() if item else ""

    def get_users_data(self):
        users_data = []
        user_map = {}
        
        for row in range(self.users_table.rowCount()):
            user_id = self._get_table_cell_text(self.users_table, row, 0)
            full_name = self._get_table_cell_text(self.users_table, row, 1)
            group = self._get_table_cell_text(self.users_table, row, 2)
            
            if not all([user_id, full_name, group]):
                continue
            
            try:
                user_id_int = int(user_id)
                if user_id_int in user_map:
                    continue
                user_map[user_id_int] = True
                
                name_parts = full_name.split()
                first_name = name_parts[1] if len(name_parts) > 1 else name_parts[0]
                last_name = name_parts[0] if name_parts else ''
                
                users_data.append((user_id_int, group, full_name, '', first_name, last_name))
            except:
                continue
        
        return users_data

    def get_schedule_data(self):
        schedule_data = []
        
        for row in range(self.schedule_table.rowCount()):
            group = self._get_table_cell_text(self.schedule_table, row, 0)
            day = self._get_table_cell_text(self.schedule_table, row, 1).lower()
            time = self._get_table_cell_text(self.schedule_table, row, 2)
            
            if group and day and time and group in GROUPS and day in DAYS_OF_WEEK:
                schedule_data.append((group, day, time))
        
        return schedule_data

    async def load_data_async(self):
        try:
            if not self.db_initialized:
                self.db = Database()
                await self.db.init_db()
                self.db_initialized = True

            users_data = self.get_users_data()
            schedule_data = self.get_schedule_data()

            if not users_data:
                return False, "Нет данных об учащихся"
            
            if not schedule_data:
                return False, "Нет данных о расписании"

            await self.db.bulk_add_users(users_data)
            await self.db.bulk_add_schedule(schedule_data)

            return True, f"Загружено {len(users_data)} учащихся и {len(schedule_data)} занятий"

        except Exception as e:
            logger.error(f"Ошибка загрузки: {e}")
            return False, str(e)

    def load_data_to_database(self):
        self.status_label.setText("Загрузка в БД...")
        self.button_add.setEnabled(False)
        
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            success, message = loop.run_until_complete(self.load_data_async())
            loop.close()
            
            if success:
                QMessageBox.information(self, "Успех", message)
                self.status_label.setText("✅ Данные загружены в БД")
                self.update_statistics()
            else:
                QMessageBox.warning(self, "Ошибка", message)
                self.status_label.setText(f"❌ {message}")
        finally:
            self.button_add.setEnabled(True)

    def save_all_data(self):
        try:
            users_data = []
            for row in range(self.users_table.rowCount()):
                row_data = [self._get_table_cell_text(self.users_table, row, col) 
                           for col in range(3)]
                if any(row_data):
                    users_data.append(row_data)
            
            with open(self.save_file_users, 'w', encoding='utf-8') as f:
                json.dump(users_data, f, ensure_ascii=False, indent=2)
            
            schedule_data = []
            for row in range(self.schedule_table.rowCount()):
                row_data = [self._get_table_cell_text(self.schedule_table, row, col) 
                           for col in range(3)]
                if any(row_data):
                    schedule_data.append(row_data)
            
            with open(self.save_file_schedule, 'w', encoding='utf-8') as f:
                json.dump(schedule_data, f, ensure_ascii=False, indent=2)
            
            self.status_label.setText("✅ Данные сохранены")
            return True
        except Exception as e:
            logger.error(f"Ошибка сохранения: {e}")
            self.status_label.setText("❌ Ошибка сохранения")
            return False

    def load_all_data(self):
        try:
            if self.save_file_users.exists():
                with open(self.save_file_users, 'r', encoding='utf-8') as f:
                    users_data = json.load(f)
                
                if users_data:
                    self.users_table.setRowCount(len(users_data))
                    self._ensure_group_widgets_for_all_rows()
                    for row, row_data in enumerate(users_data):
                        for col, value in enumerate(row_data[:3]):
                            if col == 2:
                                combo = self.users_table.cellWidget(row, col)
                                if isinstance(combo, QComboBox):
                                    combo.setCurrentText(value)
                            else:
                                self.users_table.setItem(row, col, QTableWidgetItem(str(value)))
            
            if self.save_file_schedule.exists():
                with open(self.save_file_schedule, 'r', encoding='utf-8') as f:
                    schedule_data = json.load(f)
                
                if schedule_data:
                    self.schedule_table.setRowCount(len(schedule_data))
                    self._ensure_schedule_widgets_for_all_rows()
                    for row, row_data in enumerate(schedule_data):
                        for col, value in enumerate(row_data[:3]):
                            if col in (0, 1):
                                combo = self.schedule_table.cellWidget(row, col)
                                if isinstance(combo, QComboBox):
                                    combo.setCurrentText(value)
                            else:
                                self.schedule_table.setItem(row, col, QTableWidgetItem(str(value)))
            
            self.update_statistics()
            
        except Exception as e:
            logger.error(f"Ошибка загрузки: {e}")

    def closeEvent(self, event):
        self.save_all_data()
        if self.db_initialized and self.db:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(self.db.close())
            loop.close()
        event.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
